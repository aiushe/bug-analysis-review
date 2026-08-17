#!/usr/bin/env python3
"""
build_deliverables.py — version-agnostic deliverable builder.

Takes an already-classified analysis_<version>.json (produced by
classify_headless.py, or an interactive Claude Code session following the
same schema) and renders:
  - Bug_Review_<version>_Analysis.md
  - Bug_Review_<version>_Rollup.xlsx
  - plan_<version>.json (the full write-back plan; run_review.py splits
    this into auto/review tiers via split_plan.py)

Does NO classification itself — every judgment call already happened in the
analysis JSON. All prose in the doc is either lifted verbatim from the
analysis (themes/prd_lessons/qe_lessons, if the classifier produced them) or
computed from aggregate counts — nothing here hand-authors bug-specific
narrative, which is what made the old per-release scripts unmaintainable.

USAGE
    python3 build_deliverables.py --version 1.2.6.0
"""

from __future__ import annotations

import argparse
import collections
import json
import os
import re
import sys
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
from pull_bugs import browse_base_url, load_dotenv  # type: ignore

load_dotenv(SCRIPT_DIR)
# `<site>/browse/` prefix for issue links. Empty until we know the site — the
# pulled bugs JSON records the base URL it came from, so set_browse_base() below
# recovers it even when JIRA_BASE_URL isn't exported in this shell.
BASE_BROWSE_URL = browse_base_url()


def set_browse_base(base_url: str | None) -> None:
    global BASE_BROWSE_URL
    if base_url and not BASE_BROWSE_URL:
        BASE_BROWSE_URL = f"{base_url.rstrip('/')}/browse/"

OWNER = {
    "Incomplete requirements": "PM",
    "Feature gap": "PM",
    "Requirement–expectation mismatch": "DEV",
    "Missed review by PM/QE": "QE",
    "Invalid defect": "NA",
    None: "NA",
}
OWNER_FULL = {
    "PM": "PM (product requirements)",
    "DEV": "DEV (implementation / technical spec)",
    "QE": "QE (testing, incl. cross-component)",
    "NA": "NA (invalid / unclassified)",
}
RCA_ORDER = [
    ("Requirement–expectation mismatch", "DEV"),
    ("Incomplete requirements", "PM"),
    ("Missed review by PM/QE", "QE"),
    ("Feature gap", "PM"),
    ("Invalid defect", "NA"),
]


def _clean(s: Any, n: int = 1500) -> str:
    if not s:
        return ""
    s = re.sub(r"<[^>]+>", " ", str(s))
    s = re.sub(r"\s+", " ", s).strip()
    return s[:n] + ("…" if len(s) > n else "")


def load_feature_titles(feature_pool: list[str]) -> tuple[dict[str, str], dict[str, str]]:
    """Returns (key -> title, key -> release-version-it-was-pulled-for)."""
    titles: dict[str, str] = {}
    release: dict[str, str] = {}
    for fname in feature_pool:
        path = SCRIPT_DIR / fname
        if not path.exists():
            continue
        doc = json.loads(path.read_text())
        version = (doc.get("meta") or {}).get("version", "?")
        for f in doc.get("features", []):
            key = f.get("key")
            if not key:
                continue
            titles.setdefault(key, f.get("summary") or key)
            release.setdefault(key, version)
    return titles, release


def load_bug_evidence(bug_source: str) -> dict[str, dict[str, str]]:
    path = SCRIPT_DIR / bug_source
    if not path.exists():
        return {}
    doc = json.loads(path.read_text())
    set_browse_base((doc.get("meta") or {}).get("base_url"))
    evid: dict[str, dict[str, str]] = {}
    for x in doc.get("bugs", []):
        cf = x.get("custom_fields") or {}
        evid[x["key"]] = {
            "rootcause": _clean(cf.get("customfield_10101")),
            "url": x.get("url") or f"{BASE_BROWSE_URL}{x['key']}",
        }
    return evid


def build_plan(bugs: list[dict[str, Any]], pm_rca_field: str) -> dict[str, Any]:
    """Full write-back plan — every action, gated later by split_plan.py.

    Unlike the legacy per-release scripts, an "Invalid defect" (owner NA)
    bug's RCA value IS included here (it's a writable, low-risk value) —
    the legacy scripts set a skip_reason that silently prevented it from
    ever being written, which was a latent bug, not intentional design.
    skip_reason is reserved for bugs with genuinely nothing to apply.
    """
    actions = []
    for b in sorted(bugs, key=lambda x: x["key"]):
        rca = b.get("pm_rca_type")
        feature_key = b.get("feature_key")
        if not rca and not feature_key:
            actions.append({
                "bug_key": b["key"],
                "set_pm_rca_type": None,
                "link_to_feature": None,
                "skip_reason": "no RCA classification and no feature match",
            })
            continue
        actions.append({
            "bug_key": b["key"],
            "set_pm_rca_type": rca,
            "link_to_feature": feature_key,
            "skip_reason": None,
        })
    return {
        "meta": {"version": None, "field": pm_rca_field},  # version filled by caller
        "link_type": "Requirement Link",
        "actions": actions,
    }


def render_doc(analysis: dict[str, Any], titles: dict[str, str], release: dict[str, str]) -> str:
    meta = analysis["meta"]
    bugs = analysis["bugs"]
    total = len(bugs)
    version = meta["version"]
    display = meta.get("display_version", version)

    for b in bugs:
        b["owner"] = OWNER.get(b.get("pm_rca_type"))

    owner_counts = collections.Counter(b["owner"] for b in bugs)
    rca_counts = collections.Counter(b["pm_rca_type"] for b in bugs if b.get("pm_rca_type"))
    feat_counts = collections.Counter(b["feature_key"] for b in bugs if b.get("feature_key"))
    comp_counts = collections.Counter(b["component_area"] for b in bugs if b.get("component_area"))
    conf_counts = collections.Counter(b["confidence"] for b in bugs)

    cur_release_bugs = sum(1 for b in bugs if b.get("feature_key") and release.get(b["feature_key"], "") == version)
    prior_release_bugs = sum(1 for b in bugs if b.get("feature_key") and release.get(b["feature_key"], "") not in (version, "?"))
    component_only_bugs = sum(1 for b in bugs if not b.get("feature_key") and b.get("owner") != "NA")

    L: list[str] = []

    def w(s: str = "") -> None:
        L.append(s)

    w(f"# Bug Review — Release {display} ({version})")
    w()
    w(f"*Components: {', '.join(meta.get('components', []))} · {total} bugs · "
      f"feature pool: {', '.join(meta.get('feature_pool', []))}*")
    w()
    if owner_counts:
        top_owner, top_n = owner_counts.most_common(1)[0]
        w("## Executive summary")
        w()
        w(f"- **Where the misses sit.** By owner: " +
          ", ".join(f"**{k} {owner_counts.get(k, 0)}**" for k in ("PM", "DEV", "QE", "NA")) +
          f" of {total}. Bottleneck: **{OWNER_FULL.get(top_owner, top_owner)}** "
          f"({top_n}/{total}, {round(100 * top_n / total) if total else 0}%).")
        w(f"- **Cross-release attribution.** {cur_release_bugs} bugs trace to {display} features; "
          f"{prior_release_bugs} to prior-release features; {component_only_bugs} are "
          "component-only platform regressions with no owning feature.")
        low_pct = round(100 * conf_counts.get("low", 0) / total) if total else 0
        flag = " — **exceeds the 25% documentation-gap threshold**" if low_pct > 25 else ""
        w(f"- **Confidence.** {conf_counts.get('low', 0)}/{total} bugs ({low_pct}%) low confidence{flag}.")
        w()

    w("## PM RCA Type breakdown")
    w()
    w("| PM RCA Type | Owner | Bugs |")
    w("|---|---|---:|")
    for k, o in RCA_ORDER:
        if rca_counts.get(k):
            w(f"| {k} | {o} | {rca_counts[k]} |")
    w()

    if feat_counts:
        w("## Top features by bug count")
        w()
        w("| Feature | Release | Bugs |")
        w("|---|---|---:|")
        for fk, c in feat_counts.most_common():
            w(f"| [{fk}]({BASE_BROWSE_URL}{fk}) — {titles.get(fk, fk)} | {release.get(fk, '?')} | {c} |")
        w()

    if comp_counts:
        w("## Platform components with no net-new feature (unlinked bugs)")
        w()
        w("| Component area | Bugs |")
        w("|---|---|")
        for ca, c in comp_counts.most_common():
            keys = ", ".join(f"[{b['key']}]({BASE_BROWSE_URL}{b['key']})" for b in bugs if b["component_area"] == ca)
            w(f"| {ca} | {keys} |")
        w()

    themes = analysis.get("themes") or []
    if themes:
        w("## Recurring themes")
        w()
        for t in themes:
            keys = ", ".join(t.get("bug_keys", []))
            w(f"- **{t.get('name')}** ({keys}) — {t.get('description')}")
        w()

    edge_counts = collections.Counter(b["edge_case"] for b in bugs if b.get("edge_case"))
    if edge_counts:
        w("## Top missed edge cases")
        w()
        w("| Edge case | Count |")
        w("|---|---:|")
        for e, c in edge_counts.most_common(10):
            w(f"| {e} | {c} |")
        w()

    prd_lessons = analysis.get("prd_lessons") or []
    if prd_lessons:
        w("## PRD lessons")
        for x in prd_lessons:
            w(f"- {x}")
        w()

    qe_lessons = analysis.get("qe_lessons") or []
    if qe_lessons:
        w("## QE lessons")
        for x in qe_lessons:
            w(f"- {x}")
        w()

    low_conf = [b for b in bugs if b["confidence"] == "low"]
    if low_conf:
        w("## Needs PM review (low confidence)")
        w()
        for b in low_conf:
            fc = f" → [{b['feature_key']}]({BASE_BROWSE_URL}{b['feature_key']})" if b.get("feature_key") else f" → _{b.get('component_area')}_"
            w(f"- [{b['key']}]({BASE_BROWSE_URL}{b['key']}) — {b['summary']}{fc} — {b.get('confidence_basis', '')}")
        w()

    prd_content_matches = [b for b in bugs if b.get("feature_match_basis") == "prd_content"]
    if prd_content_matches:
        w("## Feature matching — PM confirmation needed (PRD-content matches)")
        w()
        w("These were matched by reading the feature's PRD, not from an existing JIRA link. "
          "Confirm or correct before write-back:")
        w()
        for b in prd_content_matches:
            w(f"- **{b['key']}** → [{b['feature_key']}]({BASE_BROWSE_URL}{b['feature_key']}) "
              f"*{titles.get(b['feature_key'], b['feature_key'])}* — {b.get('edge_case', '')}")
        w()

    w("## Appendix — all bugs")
    w()
    w("| Bug | Summary | Owner | Feature / Component | PM RCA Type | Conf | Match basis | Edge case |")
    w("|---|---|---|---|---|---|---|---|")
    for b in sorted(bugs, key=lambda x: x["key"]):
        fc = (
            f"[{b['feature_key']}]({BASE_BROWSE_URL}{b['feature_key']}) {titles.get(b['feature_key'], '')}"
            if b.get("feature_key")
            else f"_{b.get('component_area') or ''}_"
        )
        w(f"| [{b['key']}]({BASE_BROWSE_URL}{b['key']}) | {b['summary']} | {b.get('owner', '')} | {fc} | "
          f"{b.get('pm_rca_type') or ''} | {b['confidence']} | {b.get('feature_match_basis') or ''} | "
          f"{b.get('edge_case', '')} |")
    w()

    return "\n".join(L) + "\n"


def render_xlsx(analysis: dict[str, Any], titles: dict[str, str], release: dict[str, str], evid: dict[str, Any]) -> Any:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    bugs = analysis["bugs"]
    total = len(bugs)
    for b in bugs:
        b["owner"] = OWNER.get(b.get("pm_rca_type"))

    owner_counts = collections.Counter(b["owner"] for b in bugs)
    rca_counts = collections.Counter(b["pm_rca_type"] for b in bugs if b.get("pm_rca_type"))
    feat_counts = collections.Counter(b["feature_key"] for b in bugs if b.get("feature_key"))
    comp_counts = collections.Counter(b["component_area"] for b in bugs if b.get("component_area"))
    edge_counts = collections.Counter(b["edge_case"] for b in bugs if b.get("edge_case"))
    edge_keys = collections.defaultdict(list)
    for b in bugs:
        if b.get("edge_case"):
            edge_keys[b["edge_case"]].append(b["key"])

    PINK, ZEBRA = "D61F69", "F4E7EE"
    hdr_fill = PatternFill("solid", fgColor=PINK)
    hdr_font = Font(bold=True, color="FFFFFF")
    zebra = PatternFill("solid", fgColor=ZEBRA)
    thin = Side(style="thin", color="D9C2CE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)
    wb = Workbook()

    def sheet(title: str, headers: list[str], rows: list[list[Any]], widths: list[int]) -> Any:
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(1, c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
        for i, r in enumerate(rows, start=2):
            ws.append(r)
            for c in range(1, len(headers) + 1):
                cell = ws.cell(i, c)
                cell.alignment = wrap
                cell.border = border
                if i % 2 == 0:
                    cell.fill = zebra
        for c, wd in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = wd
        ws.freeze_panes = "A2"
        return ws

    rows = []
    for b in sorted(bugs, key=lambda x: x["key"]):
        ev = evid.get(b["key"], {})
        rows.append([
            b["key"], b["summary"], b.get("owner", ""),
            b.get("feature_key") or "", titles.get(b.get("feature_key"), "") if b.get("feature_key") else "",
            b.get("component_area") or "", b.get("pm_rca_type") or "", b.get("pm_analysis") or "",
            ev.get("rootcause", ""), b["confidence"], b.get("confidence_basis", ""),
            b.get("feature_match_basis") or "", b.get("edge_case", ""),
            b.get("priority", ""), b.get("status", ""), ev.get("url", BASE_BROWSE_URL + b["key"]),
        ])
    sheet(
        "Bugs",
        ["Key", "Summary", "Owner", "Feature key", "Feature name", "Component area",
         "PM RCA Type", "PM Analysis (draft)", "Root Cause (cf_10101)", "Confidence",
         "Confidence basis", "Feature match basis", "Edge case", "Priority", "Status", "URL"],
        rows,
        [12, 40, 7, 12, 28, 22, 26, 60, 44, 10, 34, 14, 30, 16, 20, 40],
    )

    sheet("By owner", ["Owner", "Bugs", "Share %"],
          [[k, owner_counts.get(k, 0), round(100 * owner_counts.get(k, 0) / total) if total else 0]
           for k in ("PM", "DEV", "QE", "NA") if owner_counts.get(k)],
          [12, 8, 9])

    sheet("By RCA type", ["PM RCA Type", "Owner", "Bugs"],
          [[k, o, rca_counts[k]] for k, o in RCA_ORDER if rca_counts.get(k)], [40, 8, 8])

    sheet("By feature", ["Feature key", "Feature name", "Release", "Bugs"],
          [[fk, titles.get(fk, fk), release.get(fk, "?"), c] for fk, c in feat_counts.most_common()],
          [12, 42, 10, 8])

    sheet("By component", ["Component area", "Bugs", "Bug keys"],
          [[ca, c, ", ".join(b["key"] for b in bugs if b["component_area"] == ca)]
           for ca, c in comp_counts.most_common()], [30, 8, 30])

    sheet("By edge case", ["Edge case", "Count", "Bug keys"],
          [[e, c, ", ".join(edge_keys[e])] for e, c in edge_counts.most_common()], [46, 8, 40])

    wb_rows = []
    for b in sorted(bugs, key=lambda x: x["key"]):
        tier = (
            "auto" if b["confidence"] == "high" or b.get("feature_match_basis") == "linked_issues" else "review"
        )
        wb_rows.append([
            b["key"], b.get("owner", ""), b.get("pm_rca_type") or "",
            b.get("feature_key") or "", b.get("component_area") or "", tier,
        ])
    sheet("Write-back plan",
          ["Bug key", "Owner", "Set PM RCA Type", "Link to feature", "Component area", "Tier"],
          wb_rows, [12, 8, 30, 16, 24, 10])

    sheet("Meta", ["Field", "Value"],
          [["Version", analysis["meta"]["version"]],
           ["Components", ", ".join(analysis["meta"].get("components", []))],
           ["Total bugs", total],
           ["Owner split", " · ".join(f"{k} {owner_counts.get(k, 0)}" for k in ("PM", "DEV", "QE", "NA"))],
           ["Feature pool", " + ".join(analysis["meta"].get("feature_pool", []))],
           ["Classifier", analysis["meta"].get("classifier", "")]],
          [24, 90])

    del wb["Sheet"]
    return wb


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build review doc + xlsx + plan from analysis JSON.")
    p.add_argument("--version", required=True)
    p.add_argument("--out-dir", default=None)
    return p.parse_args()


def main() -> int:
    args = parse_args()
    out_dir = Path(args.out_dir) if args.out_dir else SCRIPT_DIR
    analysis_path = out_dir / f"analysis_{args.version}.json"
    analysis = json.loads(analysis_path.read_text())

    titles, release = load_feature_titles(analysis["meta"].get("feature_pool", []))
    evid = load_bug_evidence(analysis["meta"].get("bug_source", f"bugs_{args.version}.json"))

    doc_path = out_dir / f"Bug_Review_{args.version}_Analysis.md"
    doc_path.write_text(render_doc(analysis, titles, release))
    print(f"  wrote {doc_path}")

    wb = render_xlsx(analysis, titles, release, evid)
    xlsx_path = out_dir / f"Bug_Review_{args.version}_Rollup.xlsx"
    wb.save(xlsx_path)
    print(f"  wrote {xlsx_path}")

    rca_field = analysis["meta"].get("pm_rca_field") or os.environ.get(
        "JIRA_RCA_FIELD_ID", "customfield_19330"
    )
    plan = build_plan(analysis["bugs"], rca_field)
    plan["meta"]["version"] = args.version
    plan_path = out_dir / f"plan_{args.version}.json"
    plan_path.write_text(json.dumps(plan, indent=2) + "\n")
    print(f"  wrote {plan_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
